"""
YouTube 동영상 자막 요약 프로그램 (YouTube Caption Summarizer)

이 프로그램은 YouTube 동영상의 자막을 자동으로 요약해주는 도구입니다.
엑셀 파일에 있는 여러 YouTube 링크들을 일괄적으로 처리하여,
각 동영상의 자막을 가져오고 GPT를 활용해 요약합니다.

주요 기능:
1. YouTube 링크 일괄 처리
   - 엑셀 파일에서 여러 YouTube 링크를 한 번에 처리
   - 다양한 형식의 YouTube URL 지원 (일반 링크, 공유 링크, 모바일 링크 등)

2. 자막 자동 추출
   - 한국어 자막 우선 추출
   - 한국어 자막이 없는 경우 영어 자막 추출 후 번역
   - 자막이 없는 경우 적절한 오류 메시지 제공

3. GPT 기반 지능형 요약
   - GPT-4o를 활용한 고품질 요약 생성
   - 자동 번역 및 요약 통합 처리
   - 컨텍스트를 고려한 스마트 요약

4. 사용자 친화적 인터페이스
   - 직관적인 그래픽 사용자 인터페이스(GUI)
   - 진행 상황 실시간 표시
   - 자동 파일 열기 기능
   - 상세한 로그 정보 제공

5. 결과 파일 자동 포맷팅
   - 자동 줄바꿈 적용
   - 최적화된 열 너비 설정
   - 깔끔한 정렬 및 서식

사용 방법:
1. 입력 파일 준비: YouTube URL이 포함된 엑셀 파일을 준비합니다.
   (파일의 'URL' 열에 YouTube 링크가 있어야 합니다)

2. 프로그램 실행: 프로그램을 실행하고 입력 파일을 선택합니다.

3. 결과 확인: 처리가 완료되면 자동으로 결과 파일이 열립니다.

필요한 환경 설정:
1. OpenAI API 키 설정 (GPT 사용)
2. YouTube Data API 키 설정 (동영상 정보 조회)

작성자: [작성자명]
버전: 1.0.0
최종 수정일: 2024-03-23
"""

import pandas as pd
from youtube_transcript_api import YouTubeTranscriptApi
from youtube_transcript_api._errors import TranscriptsDisabled, NoTranscriptFound
from openai import OpenAI
import os
from dotenv import load_dotenv
import re
from urllib.parse import urlparse, parse_qs
import requests
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from threading import Thread
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import subprocess
import platform
import time
from datetime import datetime
import asyncio
import aiohttp

# 환경 변수 로드 및 검증
def load_api_keys():
    """API 키 로드 및 유효성 검사"""
    load_dotenv()
    api_keys = {
        'openai': os.getenv('OPENAI_API_KEY'),
        'youtube': os.getenv('YOUTUBE_API_KEY')
    }
    
    missing_keys = [key for key, value in api_keys.items() if not value]
    if missing_keys:
        raise ValueError(f"다음 API 키가 설정되지 않았습니다: {', '.join(missing_keys)}")
    
    return api_keys

def extract_video_id(url):
    """
    다양한 형식의 YouTube URL에서 비디오 ID를 추출합니다.
    
    지원하는 URL 형식:
    - https://www.youtube.com/watch?v=VIDEO_ID
    - https://youtu.be/VIDEO_ID
    - https://youtube.com/shorts/VIDEO_ID
    """
    try:
        if not url or not isinstance(url, str):
            return None
        
        # shorts URL 처리
        if 'shorts' in url:
            shorts_match = re.search(r'shorts/([a-zA-Z0-9_-]+)', url)
            if shorts_match:
                return shorts_match.group(1)
        
        # 일반적인 YouTube URL 처리
        if 'youtu.be' in url:
            return url.split('/')[-1].split('?')[0]
        elif 'youtube.com' in url:
            parsed_url = urlparse(url)
            if parsed_url.query:
                return parse_qs(parsed_url.query).get('v', [None])[0]
        
        return None
    except Exception as e:
        print(f"URL 파싱 중 오류 발생: {str(e)}")
        return None

async def get_video_title_async(video_id):
    """비동기 방식으로 비디오 제목 가져오기"""
    max_retries = 3
    retry_delay = 1
    
    for attempt in range(max_retries):
        try:
            api_keys = load_api_keys()
            youtube_api_key = api_keys['youtube']
            
            async with aiohttp.ClientSession() as session:
                url = f"https://www.googleapis.com/youtube/v3/videos?part=snippet&id={video_id}&key={youtube_api_key}"
                async with session.get(url) as response:
                    if response.status == 200:
                        data = await response.json()
                        if 'items' in data and len(data['items']) > 0:
                            return data['items'][0]['snippet']['title']
                    return "제목을 가져올 수 없습니다"
                    
        except Exception as e:
            if attempt < max_retries - 1:
                await asyncio.sleep(retry_delay)
                continue
            return f"제목 가져오기 실패: {str(e)}"

async def get_video_summary_async(video_id):
    """비동기 방식으로 비디오 요약 가져오기"""
    try:
        print(f"\n자막을 가져오는 중... (video_id: {video_id})")
        
        # 한국어 자막 우선 시도, 실패 시 영어 자막 시도
        try:
            transcript = YouTubeTranscriptApi.get_transcript(video_id, languages=['ko'])
        except NoTranscriptFound:
            try:
                transcript = YouTubeTranscriptApi.get_transcript(video_id, languages=['en'])
            except NoTranscriptFound:
                return "", "이 비디오에 대한 자막을 찾을 수 없습니다."
        
        print("자막을 성공적으로 가져왔습니다.")
        
        # 자막 텍스트 결합
        full_text = " ".join([entry['text'] for entry in transcript])
        
        # OpenAI API 키 확인
        api_keys = load_api_keys()
        client = OpenAI(api_key=api_keys['openai'])
        
        print("GPT-4o를 사용하여 요약하는 중...")
        
        # GPT를 사용하여 요약 생성
        response = await asyncio.to_thread(
            client.chat.completions.create,
            model="gpt-4o",
            messages=[
                {"role": "system", "content": "You are a helpful assistant that summarizes YouTube video transcripts. If the transcript is in Korean, summarize it in Korean. If it's in English, translate and summarize it in Korean."},
                {"role": "user", "content": f"Please summarize the following video transcript:\n\n{full_text}"}
            ],
            temperature=0.7,
            max_tokens=1000
        )
        
        print("요약이 완료되었습니다.")
        return full_text, response.choices[0].message.content
        
    except TranscriptsDisabled:
        print(f"오류: {video_id}의 스크립트가 비활성화되어 있습니다.")
        return "", "스크립트가 비활성화되어 있습니다."
    except Exception as e:
        print(f"오류 발생: {str(e)}")
        return "", f"오류가 발생했습니다: {str(e)}"

def save_excel_with_formatting(df, output_file):
    """
    데이터프레임을 엑셀 파일로 저장하고 보기 좋게 서식을 적용합니다.
    
    적용되는 서식:
    - 열 너비 자동 조정
    - 자동 줄바꿈
    - 셀 정렬
    - 헤더 스타일링
    """
    try:
        # 엑셀 파일로 저장
        df.to_excel(output_file, index=False)
        
        # openpyxl로 파일을 열어서 서식 적용
        wb = load_workbook(output_file)
        ws = wb.active
        
        # 열 너비 설정
        column_widths = {
            '제목': 40,
            'URL': 30,
            '원본 자막': 60,
            'GPT 요약': 60
        }
        
        # 모든 열에 대해 서식 적용
        for idx, col in enumerate(df.columns, 1):
            column_letter = get_column_letter(idx)
            
            # 열 너비 설정
            ws.column_dimensions[column_letter].width = column_widths.get(col, 30)
            
            # 첫 번째 행(헤더) 서식 설정
            header_cell = ws[f"{column_letter}1"]
            header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # 데이터 행 서식 설정
            for row in range(2, len(df) + 2):
                cell = ws[f"{column_letter}{row}"]
                
                # 원본 자막과 GPT 요약 열에만 자동 줄 바꿈 적용
                if col in ['원본 자막', 'GPT 요약']:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical='top')
        
        # 변경사항 저장
        wb.save(output_file)
    except Exception as e:
        raise Exception(f"엑셀 파일 저장 중 오류 발생: {str(e)}")

def open_file(filepath):
    """
    운영체제에 따라 적절한 방법으로 파일을 엽니다.
    재시도 로직이 포함되어 있습니다.
    """
    max_retries = 3
    retry_delay = 1  # 초
    
    for attempt in range(max_retries):
        try:
            if platform.system() == 'Windows':
                os.startfile(filepath)
            elif platform.system() == 'Darwin':  # macOS
                subprocess.run(['open', filepath])
            else:  # Linux
                subprocess.run(['xdg-open', filepath])
            return True
        except Exception as e:
            if attempt < max_retries - 1:
                time.sleep(retry_delay)
                continue
            print(f"파일을 열 수 없습니다: {str(e)}")
            return False

class YouTubeSummarizerGUI:
    """
    YouTube 동영상 자막 요약 프로그램의 그래픽 사용자 인터페이스
    """
    def __init__(self, root):
        """GUI 초기화 및 구성요소 배치"""
        self.root = root
        self.root.title("YouTube 비디오 요약 도구")
        self.root.geometry("650x700")  # 창 크기 증가
        
        # 종료 프로토콜 설정
        self.root.protocol("WM_DELETE_WINDOW", self.exit_program)
        
        # 스타일 설정
        self.setup_styles()
        
        # 메인 프레임 설정
        self.setup_main_frame()
        
        # 노트북(탭) 생성
        self.notebook = ttk.Notebook(self.main_frame)
        self.notebook.grid(row=0, column=0, sticky='nsew', padx=5, pady=5)
        
        # URL 직접 입력 탭 (먼저 추가)
        self.url_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.url_frame, text='URL 직접 입력')
        
        # 엑셀 처리 탭
        self.excel_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.excel_frame, text='엑셀 파일 처리')
        
        # 각 탭 구성
        self.setup_url_tab()  # URL 탭 먼저 설정
        self.setup_excel_tab()
        
        # 공통 UI 구성
        self.setup_common_ui()
        
        # 처리 상태 표시기
        self.processing = False
        self.max_retries = 3
        self.start_time = None

    def setup_styles(self):
        """스타일 설정"""
        style = ttk.Style()
        style.configure("Custom.TButton", padding=5)
        style.configure("Custom.TLabel", padding=5)
        style.configure("Title.TLabel", font=('Helvetica', 12, 'bold'))

    def setup_main_frame(self):
        """메인 프레임 설정"""
        self.main_frame = ttk.Frame(self.root, padding="10")
        self.main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

    def setup_excel_tab(self):
        """엑셀 파일 처리 탭 구성"""
        # 입력 파일 선택
        ttk.Label(self.excel_frame, text="입력 엑셀 파일:", style="Custom.TLabel").grid(row=0, column=0, sticky=tk.W)
        self.input_path_var = tk.StringVar()
        self.input_path_entry = ttk.Entry(self.excel_frame, textvariable=self.input_path_var, width=50)
        self.input_path_entry.grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(self.excel_frame, text="찾아보기", command=self.select_input_file, style="Custom.TButton").grid(row=0, column=2)
        
        # 출력 파일 선택
        ttk.Label(self.excel_frame, text="출력 엑셀 파일:", style="Custom.TLabel").grid(row=1, column=0, sticky=tk.W)
        self.output_path_var = tk.StringVar()
        self.output_path_entry = ttk.Entry(self.excel_frame, textvariable=self.output_path_var, width=50)
        self.output_path_entry.grid(row=1, column=1, padx=5, pady=5)
        ttk.Button(self.excel_frame, text="찾아보기", command=self.select_output_file, style="Custom.TButton").grid(row=1, column=2)
        
        # 버튼 프레임 생성
        button_frame = ttk.Frame(self.excel_frame)
        button_frame.grid(row=2, column=1, pady=10)
        
        # 엑셀 처리 시작 버튼
        self.excel_start_button = ttk.Button(button_frame, text="엑셀 처리 시작", 
                                           command=lambda: self.start_processing('excel'), 
                                           style="Custom.TButton")
        self.excel_start_button.pack(side=tk.LEFT, padx=5)
        
        # 종료 버튼
        self.excel_stop_button = ttk.Button(button_frame, text="처리 중단", 
                                          command=self.stop_processing,
                                          style="Custom.TButton",
                                          state='disabled')
        self.excel_stop_button.pack(side=tk.LEFT, padx=5)

    def setup_url_tab(self):
        """URL 직접 입력 탭 구성"""
        # URL 입력
        ttk.Label(self.url_frame, text="YouTube URL:", style="Custom.TLabel").grid(row=0, column=0, sticky=tk.W)
        self.url_var = tk.StringVar()
        self.url_entry = ttk.Entry(self.url_frame, textvariable=self.url_var, width=70)
        self.url_entry.grid(row=0, column=1, padx=5, pady=5, columnspan=2)
        
        # 출력 파일 선택
        ttk.Label(self.url_frame, text="저장할 파일명:", style="Custom.TLabel").grid(row=1, column=0, sticky=tk.W)
        self.url_output_var = tk.StringVar()
        self.url_output_entry = ttk.Entry(self.url_frame, textvariable=self.url_output_var, width=50)
        self.url_output_entry.grid(row=1, column=1, padx=5, pady=5)
        ttk.Button(self.url_frame, text="찾아보기", command=self.select_url_output_file, style="Custom.TButton").grid(row=1, column=2)
        
        # 버튼 프레임 생성
        button_frame = ttk.Frame(self.url_frame)
        button_frame.grid(row=2, column=1, pady=10)
        
        # URL 처리 시작 버튼
        self.url_start_button = ttk.Button(button_frame, text="URL 처리 시작", 
                                         command=lambda: self.start_processing('url'), 
                                         style="Custom.TButton")
        self.url_start_button.pack(side=tk.LEFT, padx=5)
        
        # 종료 버튼
        self.url_stop_button = ttk.Button(button_frame, text="처리 중단", 
                                        command=self.stop_processing,
                                        style="Custom.TButton",
                                        state='disabled')
        self.url_stop_button.pack(side=tk.LEFT, padx=5)
        
        # 결과 표시 영역
        self.result_text = tk.Text(self.url_frame, height=15, width=80)
        self.result_text.grid(row=3, column=0, columnspan=3, pady=10)
        
        # 스크롤바 추가
        result_scrollbar = ttk.Scrollbar(self.url_frame, orient="vertical", command=self.result_text.yview)
        result_scrollbar.grid(row=3, column=3, sticky="ns")
        self.result_text.configure(yscrollcommand=result_scrollbar.set)

    def setup_common_ui(self):
        """공통 UI 구성"""
        common_frame = ttk.Frame(self.main_frame)
        common_frame.grid(row=1, column=0, sticky='nsew', padx=5, pady=5)
        
        # 자동 열기 체크박스
        self.auto_open_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(common_frame, text="완료 후 결과 파일 자동으로 열기", 
                       variable=self.auto_open_var).grid(row=0, column=0, columnspan=3, pady=5)
        
        # 진행 상황 표시
        self.progress_var = tk.DoubleVar()
        self.progress = ttk.Progressbar(common_frame, length=600, mode='determinate', variable=self.progress_var)
        self.progress.grid(row=1, column=0, columnspan=3, pady=10)
        
        # 상태 메시지
        self.status_var = tk.StringVar(value="URL을 입력하거나 엑셀 파일을 선택해주세요.")
        self.status_label = ttk.Label(common_frame, textvariable=self.status_var, wraplength=700)
        self.status_label.grid(row=2, column=0, columnspan=3)
        
        # 로그 창
        self.log_text = tk.Text(common_frame, height=10, width=80)
        self.log_text.grid(row=3, column=0, columnspan=3, pady=10)
        
        # 로그 스크롤바
        log_scrollbar = ttk.Scrollbar(common_frame, orient="vertical", command=self.log_text.yview)
        log_scrollbar.grid(row=3, column=3, sticky="ns")
        self.log_text.configure(yscrollcommand=log_scrollbar.set)
        
        # 하단 버튼 프레임
        bottom_frame = ttk.Frame(common_frame)
        bottom_frame.grid(row=4, column=0, columnspan=3, pady=10)
        
        # 프로그램 종료 버튼
        self.exit_button = ttk.Button(bottom_frame, 
                                    text="프로그램 종료", 
                                    command=self.exit_program,
                                    style="Custom.TButton")
        self.exit_button.pack(pady=5)

    def exit_program(self):
        """프로그램 종료"""
        if self.processing:
            if not messagebox.askyesno("확인", 
                                     "처리가 진행 중입니다.\n정말로 프로그램을 종료하시겠습니까?"):
                return
            self.processing = False
            self.log_message("사용자가 프로그램을 종료했습니다.")
        
        self.root.quit()
        self.root.destroy()

    def select_input_file(self):
        """입력 파일 선택 다이얼로그"""
        filename = filedialog.askopenfilename(
            title="입력 엑셀 파일 선택",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.input_path_var.set(filename)
            if not self.output_path_var.get():
                # 기본 출력 파일명 생성 (입력파일명_요약결과_날짜시간.xlsx)
                base_name = os.path.splitext(filename)[0]
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_filename = f"{base_name}_요약결과_{timestamp}.xlsx"
                self.output_path_var.set(output_filename)

    def select_output_file(self):
        """출력 파일 선택 다이얼로그"""
        filename = filedialog.asksaveasfilename(
            title="출력 엑셀 파일 선택",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            defaultextension=".xlsx"
        )
        if filename:
            self.output_path_var.set(filename)

    def select_url_output_file(self):
        """URL 처리 결과 저장 파일 선택 다이얼로그"""
        filename = filedialog.asksaveasfilename(
            title="결과 저장 파일 선택",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            defaultextension=".xlsx"
        )
        if filename:
            self.url_output_var.set(filename)

    def log_message(self, message):
        """로그 메시지 추가"""
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()

    def update_status(self, message):
        """상태 메시지 업데이트"""
        self.status_var.set(message)
        self.root.update_idletasks()

    def process_single_url(self, url):
        """단일 URL 처리"""
        try:
            video_id = extract_video_id(url)
            if not video_id:
                return "유효하지 않은 YouTube URL입니다."
            
            self.log_message(f"Processing URL: {url}")
            
            # 비동기 함수를 동기적으로 실행
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            title = loop.run_until_complete(get_video_title_async(video_id))
            original_text, summary = loop.run_until_complete(get_video_summary_async(video_id))
            loop.close()
            
            result = f"제목: {title}\n\n"
            result += f"URL: {url}\n\n"
            result += f"원본 자막:\n{original_text}\n\n"
            result += f"GPT 요약:\n{summary}\n"
            
            # 결과를 엑셀 파일로 저장
            output_file = self.url_output_var.get()
            if not output_file:
                # 기본 출력 파일명 생성 (입력파일명_요약결과_날짜시간.xlsx)
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_file = f"youtube_summary_{timestamp}.xlsx"
                self.url_output_var.set(output_file)
            
            df = pd.DataFrame([{
                '제목': title,
                'URL': url,
                '원본 자막': original_text,
                'GPT 요약': summary
            }])
            
            save_excel_with_formatting(df, output_file)
            
            if self.auto_open_var.get():
                self.root.after(1000, lambda: open_file(output_file))
            
            return result
            
        except Exception as e:
            return f"처리 중 오류가 발생했습니다: {str(e)}"

    def stop_processing(self):
        """처리 중단"""
        if self.processing:
            self.processing = False
            self.update_status("처리가 중단되었습니다.")
            self.log_message("사용자가 처리를 중단했습니다.")
            
            # 현재 활성화된 탭에 따라 적절한 버튼 상태 변경
            current_tab = self.notebook.index(self.notebook.select())
            if current_tab == 0:  # URL 탭
                self.url_start_button.state(['!disabled'])
                self.url_stop_button.state(['disabled'])
            else:  # 엑셀 탭
                self.excel_start_button.state(['!disabled'])
                self.excel_stop_button.state(['disabled'])

    def start_processing(self, mode='excel'):
        """처리 시작"""
        if self.processing:
            return
            
        try:
            # API 키 검증
            load_api_keys()
            
            if mode == 'excel':
                input_file = self.input_path_var.get()
                output_file = self.output_path_var.get()
                
                if not input_file or not output_file:
                    messagebox.showerror("오류", "입력 파일과 출력 파일을 모두 선택해주세요.")
                    return
                    
                self.processing = True
                self.excel_start_button.state(['disabled'])
                self.excel_stop_button.state(['!disabled'])
                self.progress_var.set(0)
                self.start_time = datetime.now()
                
                # 엑셀 처리 스레드 시작
                thread = Thread(target=self.process_excel_thread, args=(input_file, output_file))
                thread.daemon = True
                thread.start()
                
            else:  # URL 모드
                url = self.url_var.get().strip()
                if not url:
                    messagebox.showerror("오류", "YouTube URL을 입력해주세요.")
                    return
                    
                self.processing = True
                self.url_start_button.state(['disabled'])
                self.url_stop_button.state(['!disabled'])
                self.progress_var.set(0)
                self.start_time = datetime.now()
                
                # URL 처리 스레드 시작
                thread = Thread(target=self.process_url_thread, args=(url,))
                thread.daemon = True
                thread.start()
                
        except ValueError as e:
            messagebox.showerror("API 키 오류", str(e))
        except Exception as e:
            messagebox.showerror("오류", f"처리 시작 중 오류가 발생했습니다:\n{str(e)}")

    def process_url_thread(self, url):
        """URL 처리 스레드"""
        try:
            self.progress_var.set(20)
            self.update_status("URL 처리 중...")
            
            result = self.process_single_url(url)
            
            self.progress_var.set(100)
            self.result_text.delete(1.0, tk.END)
            self.result_text.insert(tk.END, result)
            
            self.update_status("URL 처리가 완료되었습니다!")
            
        except Exception as e:
            self.update_status(f"오류가 발생했습니다: {str(e)}")
            self.log_message(f"심각한 오류 발생: {str(e)}")
            messagebox.showerror("오류", f"처리 중 오류가 발생했습니다:\n{str(e)}")
            
        finally:
            self.processing = False
            self.url_start_button.state(['!disabled'])

    def process_excel_thread(self, input_file, output_file):
        """엑셀 파일 처리 (별도 스레드)"""
        try:
            # 엑셀 파일 읽기
            df = pd.read_excel(input_file)
            total_rows = len(df)
            results = []
            
            # 이벤트 루프 생성
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            
            for index, row in df.iterrows():
                # 처리 중단 확인
                if not self.processing:
                    self.log_message("처리가 중단되었습니다.")
                    return
                    
                url = row['URL']
                video_id = extract_video_id(url)
                progress = (index + 1) / total_rows * 100
                
                self.progress_var.set(progress)
                self.update_status(f"처리 중... ({index + 1}/{total_rows})")
                
                if video_id:
                    self.log_message(f"Processing video {index + 1}: {url}")
                    title = loop.run_until_complete(get_video_title_async(video_id))
                    original_text, summary = loop.run_until_complete(get_video_summary_async(video_id))
                    results.append({
                        '제목': title,
                        'URL': url,
                        '원본 자막': original_text,
                        'GPT 요약': summary
                    })
                else:
                    self.log_message(f"잘못된 URL 형식: {url}")
                    results.append({
                        '제목': '유효하지 않은 URL',
                        'URL': url,
                        '원본 자막': '',
                        'GPT 요약': '유효하지 않은 YouTube URL입니다.'
                    })
            
            # 이벤트 루프 종료
            loop.close()
            
            # 결과를 DataFrame으로 변환하고 저장
            result_df = pd.DataFrame(results)
            
            # 엑셀 파일 저장 및 서식 적용
            save_excel_with_formatting(result_df, output_file)
            
            # 처리 시간 계산
            end_time = datetime.now()
            processing_time = end_time - self.start_time
            
            self.update_status("처리가 완료되었습니다!")
            self.log_message(f"총 처리 시간: {processing_time}")
            messagebox.showinfo("완료", f"요약이 완료되었습니다.\n결과가 {output_file}에 저장되었습니다.\n총 처리 시간: {processing_time}")
            
            # 자동 열기가 체크되어 있으면 파일 열기
            if self.auto_open_var.get():
                self.root.after(1000, lambda: open_file(output_file))
            
        except Exception as e:
            self.update_status(f"오류가 발생했습니다: {str(e)}")
            self.log_message(f"심각한 오류 발생: {str(e)}")
            messagebox.showerror("오류", f"처리 중 오류가 발생했습니다:\n{str(e)}")
            
        finally:
            self.processing = False
            self.excel_start_button.state(['!disabled'])
            self.excel_stop_button.state(['disabled'])

def main():
    """메인 함수"""
    try:
        root = tk.Tk()
        app = YouTubeSummarizerGUI(root)
        root.mainloop()
    except Exception as e:
        messagebox.showerror("심각한 오류", f"프로그램 실행 중 오류가 발생했습니다:\n{str(e)}")

if __name__ == "__main__":
    main() 